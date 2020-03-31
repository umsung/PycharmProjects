import pymysql as pms
from datetime import datetime,timedelta
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib

class SqlExc(object):
    def __init__(self,host='',port='',username='',password='',database=''):
        self.conn = pms.connect(host=host,port=port,user=username,passwd=password,database=database,charset='utf-8')
        self.cur = self.conn.cursor()
        self.sql1=''
        self.sql2=''
        self.email_from=''
        self.email_to=''
        self.email_subject=''

    def getDataAndColumn(self):
        datas = self.cur.excute(self.sql1).fetchall()
        fields = self.cur.excute(self.sql2).description()
        return datas,fields

    def get_excel(self,datas,fields,filePath):
        # 创建一个表对象
        new = openpyxl.Workbook()
        # 创建sheet
        sheet = new.active()
        sheet.title = 'sheet名'
        # 将字段名称循环写入excel第一行，因为字段格式列表里包含列表，每个列表的第一元素才是字段名称
        for i in range(len(fields)):
            #  cow代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
            sheet.cell(row=1,column=i+1,value=fields[i][0])

        for row in range(1,len(datas)+1):
            for col in range(len(fields)):
                sheet.cell(row=row+1,column=col+1,value=datas[row-1][col])
        newWorkbook = new.save(filePath)
        return newWorkbook

    def getYesterday(self):
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        yesterdayStr = yesterday.strftime('%Y-%m-%d')
        return yesterdayStr

    def create_email(self,email_text,filename,filepath):
        # 生成带附件的邮件实例
        message = MIMEMultipart()
        # 将正文以text的形式插入邮件中
        message.attach(MIMEText(email_text,'palin','utf-8'))
        message['From'] = Header(self.email_from,'utf-8')  # 生成发件人名称（这不是接收的邮件）
        message['to'] = Header(self.email_to,'utf-8')  # 生成收件人名称（这不是接收的邮件）
        message['Subject'] = Header(self.email_subject,'utf-8')

        # 读取带附件的内容
        att1 = MIMEText(open(filepath,'rb'),'base64','utf-8')
        att1["Content-Type"] = 'application/octet-stream'

        # 生成附件的名称
        att1["Content-Disposition"] = 'attachment; filename=' + filename

        # 将附件插入邮箱
        message.attach(att1)
        return message

    def send_email(self,mail_host, mail_port,mail_user, mail_pwd,message,sender,receiver):
        try:
            # 实例化SMTP()
            smtp_obj = smtplib.SMTP()

            # 实例化SMTP(),连接邮箱服务器, mail_host是邮箱服务器地址，mail_port是端口，
            # 新浪邮箱：smtp.sina.com,
            # 新浪VIP：smtp.vip.sina.com,
            # 搜狐邮箱：smtp.sohu.com，
            # 126邮箱：smtp.126.com,
            # 139邮箱：smtp.139.com,
            # 163网易邮箱：smtp.163.com。
            smtp_obj.connect(mail_host, mail_port)  # SMTP协议默认端口是25
            smtp_obj.login(mail_user, mail_pwd)
            # as_string()message(MIMEText对象或者MIMEMultipart对象)变为str
            smtp_obj.sendmail(sender, receiver, message.as_string())
            smtp_obj.quit()
        except smtplib.SMTPException as e:
            # print(traceback.print_exc())
            print("邮件发送失败")